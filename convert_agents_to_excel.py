import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def parse_agents_file(filepath):
    """Parse the agents.md file and extract agent information."""
    agents = []
    
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Split by lines starting with "-"
    lines = content.strip().split('\n')
    
    for line in lines:
        line = line.strip()
        if not line or not line.startswith('-'):
            continue
            
        # Remove the leading "- "
        line = line[2:].strip()
        
        # Parse the line using regex or split by |
        parts = line.split('|')
        
        if len(parts) >= 4:
            name = parts[0].strip()
            
            # Extract LICENCE
            licence = ''
            if 'LICENCE:' in parts[1]:
                licence = parts[1].split('LICENCE:')[1].strip()
            
            # Extract MOBILE
            mobile = ''
            if 'MOBILE:' in parts[2]:
                mobile = parts[2].split('MOBILE:')[1].strip()
            
            # Extract HOME
            home = ''
            if len(parts) > 3 and 'HOME:' in parts[3]:
                home = parts[3].split('HOME:')[1].strip()
            
            # Extract EMAIL
            email = ''
            if len(parts) > 4 and 'EMAIL:' in parts[4]:
                email = parts[4].split('EMAIL:')[1].strip()
            
            agents.append({
                'Name': name,
                'Licence': licence,
                'Mobile': mobile,
                'Home': home,
                'Email': email
            })
    
    return agents

def create_excel(agents, output_file):
    """Create an Excel file from the agents data."""
    # Create DataFrame
    df = pd.DataFrame(agents)
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Agents"
    
    # Write data to worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Style header row
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40  # Name
    ws.column_dimensions['B'].width = 15  # Licence
    ws.column_dimensions['C'].width = 25  # Mobile
    ws.column_dimensions['D'].width = 25  # Home
    ws.column_dimensions['E'].width = 35  # Email
    
    # Save the workbook
    wb.save(output_file)
    print(f"Excel file created successfully: {output_file}")
    print(f"Total agents: {len(agents)}")

if __name__ == "__main__":
    input_file = "agents.md"
    output_file = "agents.xlsx"
    
    print("Parsing agents.md file...")
    agents = parse_agents_file(input_file)
    
    print(f"Found {len(agents)} agents")
    
    print("Creating Excel file...")
    create_excel(agents, output_file)
    
    print("\nFirst 5 agents:")
    for i, agent in enumerate(agents[:5], 1):
        print(f"{i}. {agent['Name']} - {agent['Licence']}")
